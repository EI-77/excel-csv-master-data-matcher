import json
import math
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_DIR = "input"
MASTER_DIR = "master"
OUTPUT_DIR = "output"
CONFIG_FILE = "config.json"

SUPPORTED_INPUT_EXTENSIONS = [".csv", ".xlsx"]
EXCEL_EXTENSION = ".xlsx"
CSV_EXTENSION = ".csv"

DEFAULT_CSV_READ_ENCODINGS = ["utf-8-sig", "utf-8", "cp932"]
DEFAULT_DATE_FORMAT = "%Y-%m-%d"

INTERNAL_ROW_NUMBER = "__row_number"
INTERNAL_MASTER_ROW_NUMBER = "__master_row_number"
MERGE_STATUS_COLUMN = "_merge"


# ------------------------------------------------------------
# Sample file creation
# ------------------------------------------------------------

def create_sample_files():
    """
    Create folders, config.json, sample order data, and sample master data.

    Existing files are not overwritten.
    """
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(MASTER_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.exists(CONFIG_FILE):
        sample_config = {
            "order_file": "sample_orders.xlsx",
            "master_file": "product_master.xlsx",
            "enriched_file": "enriched_orders.xlsx",
            "unmatched_file": "unmatched_orders.xlsx",
            "summary_file": "match_summary.xlsx",
            "order_key": "product_code",
            "master_key": "product_code",
            "order_required_columns": [
                "order_id",
                "order_date",
                "product_code",
                "quantity"
            ],
            "master_required_columns": [
                "product_code",
                "product_name",
                "category",
                "standard_price"
            ],
            "master_value_columns": [
                "product_name",
                "category",
                "standard_price"
            ],
            "order_numeric_columns": [
                "quantity"
            ],
            "master_numeric_columns": [
                "standard_price"
            ],
            "order_date_columns": [
                "order_date"
            ],
            "date_format": "%Y-%m-%d",
            "amount_calculation": {
                "enabled": True,
                "quantity_column": "quantity",
                "price_column": "standard_price",
                "output_column": "expected_amount"
            },
            "enriched_output_columns": [
                "order_id",
                "order_date",
                "product_code",
                "product_name",
                "category",
                "quantity",
                "standard_price",
                "expected_amount"
            ],
            "unmatched_output_columns": [
                "row_number",
                "order_id",
                "order_date",
                "product_code",
                "quantity",
                "error_type",
                "error_message"
            ]
        }

        with open(CONFIG_FILE, "w", encoding="utf-8") as file:
            json.dump(sample_config, file, ensure_ascii=False, indent=2)

        print(f"Sample config created: {CONFIG_FILE}")

    sample_order_path = os.path.join(INPUT_DIR, "sample_orders.xlsx")
    sample_master_path = os.path.join(MASTER_DIR, "product_master.xlsx")

    if not os.path.exists(sample_order_path):
        create_sample_orders_excel(sample_order_path)
        print(f"Sample order file created: {sample_order_path}")

    if not os.path.exists(sample_master_path):
        create_product_master_excel(sample_master_path)
        print(f"Sample master file created: {sample_master_path}")


def create_sample_orders_excel(file_path):
    """Create sample order data with matched rows and intentional errors."""
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"

    headers = [
        "order_id",
        "order_date",
        "product_code",
        "quantity",
        "note"
    ]

    rows = [
        ["ORD-001", "2026/06/12", "P-1001", 2, "正常データ"],
        ["ORD-002", "2026/06/13", "P-1002", 1, "正常データ"],
        ["ORD-003", "2026/06/14", "P-9999", 3, "マスタに存在しない商品コード"],
        ["ORD-004", "2026/06/15", "", 1, "商品コードが空欄"],
        ["ORD-005", "2026/06/16", "P-1003", "abc", "数量が数値ではない"],
        ["ORD-006", "2026/06/17", "P-1004", 2, "正常データ"]
    ]

    ws.append(headers)

    for row in rows:
        ws.append(row)

    apply_basic_worksheet_format(ws)
    wb.save(file_path)


def create_product_master_excel(file_path):
    """Create sample product master data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "product_master"

    headers = [
        "product_code",
        "product_name",
        "category",
        "standard_price"
    ]

    rows = [
        ["P-1001", "ワイヤレスマウス", "PC周辺機器", 1800],
        ["P-1002", "USBキーボード", "PC周辺機器", 3200],
        ["P-1003", "Webカメラ", "PC周辺機器", 4500],
        ["P-1004", "外付けSSD", "ストレージ", 9800]
    ]

    ws.append(headers)

    for row in rows:
        ws.append(row)

    apply_basic_worksheet_format(ws)
    wb.save(file_path)


# ------------------------------------------------------------
# Config handling
# ------------------------------------------------------------

def load_config():
    """Load config.json as a dictionary."""
    with open(CONFIG_FILE, "r", encoding="utf-8") as file:
        return json.load(file)


def validate_config(config):
    """Validate config.json before reading input files."""
    required_keys = [
        "order_file",
        "master_file",
        "enriched_file",
        "unmatched_file",
        "summary_file",
        "order_key",
        "master_key",
        "order_required_columns",
        "master_required_columns",
        "master_value_columns",
        "order_numeric_columns",
        "master_numeric_columns",
        "order_date_columns",
        "enriched_output_columns",
        "unmatched_output_columns"
    ]

    missing_keys = [key for key in required_keys if key not in config]

    if missing_keys:
        raise ValueError(f"Missing config keys: {', '.join(missing_keys)}")

    for key in [
        "order_file",
        "master_file",
        "enriched_file",
        "unmatched_file",
        "summary_file",
        "order_key",
        "master_key"
    ]:
        validate_string_value(config, key)

    for key in [
        "order_required_columns",
        "master_required_columns",
        "master_value_columns",
        "order_numeric_columns",
        "master_numeric_columns",
        "order_date_columns",
        "enriched_output_columns",
        "unmatched_output_columns"
    ]:
        validate_string_list(config, key)

    if "date_format" in config:
        validate_string_value(config, "date_format")

    validate_file_extension(config["order_file"], SUPPORTED_INPUT_EXTENSIONS, "order_file")
    validate_file_extension(config["master_file"], SUPPORTED_INPUT_EXTENSIONS, "master_file")
    validate_file_extension(config["enriched_file"], [EXCEL_EXTENSION], "enriched_file")
    validate_file_extension(config["unmatched_file"], [EXCEL_EXTENSION], "unmatched_file")
    validate_file_extension(config["summary_file"], [EXCEL_EXTENSION], "summary_file")

    if config["order_key"] not in config["order_required_columns"]:
        raise ValueError("order_key must be included in order_required_columns.")

    if config["master_key"] not in config["master_required_columns"]:
        raise ValueError("master_key must be included in master_required_columns.")

    for column in config["master_value_columns"]:
        if column == config["master_key"]:
            raise ValueError("master_value_columns must not include master_key.")

    validate_amount_calculation(config)


def validate_amount_calculation(config):
    """Validate optional amount calculation settings."""
    amount_config = config.get("amount_calculation", {"enabled": False})

    if not isinstance(amount_config, dict):
        raise ValueError("amount_calculation must be an object.")

    if not amount_config.get("enabled", False):
        return

    required_keys = ["quantity_column", "price_column", "output_column"]
    missing_keys = [key for key in required_keys if key not in amount_config]

    if missing_keys:
        raise ValueError(f"Missing amount_calculation keys: {', '.join(missing_keys)}")

    for key in required_keys:
        if not isinstance(amount_config[key], str) or amount_config[key].strip() == "":
            raise ValueError(f"amount_calculation.{key} must be a non-empty string.")

    quantity_column = amount_config["quantity_column"]
    price_column = amount_config["price_column"]
    output_column = amount_config["output_column"]

    if quantity_column not in config["order_numeric_columns"]:
        raise ValueError("amount_calculation.quantity_column must be included in order_numeric_columns.")

    if price_column not in config["master_numeric_columns"]:
        raise ValueError("amount_calculation.price_column must be included in master_numeric_columns.")

    if output_column not in config["enriched_output_columns"]:
        raise ValueError("amount_calculation.output_column must be included in enriched_output_columns.")


def validate_string_value(config, key):
    """Validate that config[key] is a non-empty string."""
    if not isinstance(config[key], str) or config[key].strip() == "":
        raise ValueError(f"{key} must be a non-empty string.")


def validate_string_list(config, key):
    """Validate that config[key] is a list of non-empty strings."""
    if not isinstance(config[key], list):
        raise ValueError(f"{key} must be a list.")

    invalid_values = [
        value for value in config[key]
        if not isinstance(value, str) or value.strip() == ""
    ]

    if invalid_values:
        raise ValueError(f"{key} must contain only non-empty strings.")


def validate_file_extension(file_name, allowed_extensions, config_key):
    """Validate a file extension against allowed extensions."""
    extension = os.path.splitext(file_name)[1].lower()

    if extension not in allowed_extensions:
        allowed_text = ", ".join(allowed_extensions)
        raise ValueError(f"{config_key} must be one of these types: {allowed_text}")


# ------------------------------------------------------------
# Path helpers
# ------------------------------------------------------------

def make_order_path(config):
    """Return the order file path under input directory."""
    return os.path.join(INPUT_DIR, config["order_file"])


def make_master_path(config):
    """Return the master file path under master directory."""
    return os.path.join(MASTER_DIR, config["master_file"])


def make_enriched_path(config):
    """Return the enriched output file path under output directory."""
    return os.path.join(OUTPUT_DIR, config["enriched_file"])


def make_unmatched_path(config):
    """Return the unmatched output file path under output directory."""
    return os.path.join(OUTPUT_DIR, config["unmatched_file"])


def make_summary_path(config):
    """Return the summary output file path under output directory."""
    return os.path.join(OUTPUT_DIR, config["summary_file"])


# ------------------------------------------------------------
# Input reading
# ------------------------------------------------------------

def read_table_file(file_path):
    """Read a CSV or Excel file as a DataFrame."""
    extension = os.path.splitext(file_path)[1].lower()

    if extension == CSV_EXTENSION:
        df = read_csv_with_fallback_encodings(file_path)
    elif extension == EXCEL_EXTENSION:
        df = pd.read_excel(file_path, dtype=object)
    else:
        raise ValueError(f"Unsupported file type: {file_path}")

    return normalize_column_names(df, file_path)


def read_csv_with_fallback_encodings(file_path):
    """Read a CSV file by trying common Japanese encodings."""
    last_error = None

    for encoding in DEFAULT_CSV_READ_ENCODINGS:
        try:
            return pd.read_csv(file_path, dtype=object, encoding=encoding)
        except UnicodeDecodeError as error:
            last_error = error

    raise UnicodeDecodeError(
        last_error.encoding,
        last_error.object,
        last_error.start,
        last_error.end,
        f"Failed to read CSV with encodings: {', '.join(DEFAULT_CSV_READ_ENCODINGS)}"
    )


def normalize_column_names(df, file_path):
    """Strip column names and reject duplicated column names."""
    df = df.copy()
    df.columns = [str(column).strip() for column in df.columns]

    duplicated_columns = [
        column for column in df.columns
        if list(df.columns).count(column) > 1
    ]

    if duplicated_columns:
        unique_duplicated_columns = sorted(set(duplicated_columns))
        raise ValueError(
            f"Duplicated columns found in {file_path}: "
            f"{', '.join(unique_duplicated_columns)}"
        )

    return df


# ------------------------------------------------------------
# Validation and normalization
# ------------------------------------------------------------

def check_required_columns(df, required_columns, file_label):
    """Ensure all required columns exist in a DataFrame."""
    missing_columns = [
        column for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            f"Missing columns in {file_label}: {', '.join(missing_columns)}"
        )


def normalize_dataframe_values(df):
    """Normalize all visible cell values in a DataFrame."""
    normalized_df = df.copy()
    return normalized_df.apply(lambda column: column.map(normalize_general_value))


def prepare_order_dataframe(order_df):
    """Normalize order values and add original row numbers."""
    prepared_df = normalize_dataframe_values(order_df)
    prepared_df[INTERNAL_ROW_NUMBER] = list(range(2, len(prepared_df) + 2))
    return prepared_df


def prepare_master_dataframe(master_df):
    """Normalize master values and add original row numbers."""
    prepared_df = normalize_dataframe_values(master_df)
    prepared_df[INTERNAL_MASTER_ROW_NUMBER] = list(range(2, len(prepared_df) + 2))
    return prepared_df


def validate_and_prepare_master(master_df, config):
    """
    Validate master data as baseline data.

    Master errors stop the whole process because the master file is used as
    the source of truth for matching and enrichment.
    """
    master_key = config["master_key"]
    required_columns = config["master_required_columns"]
    numeric_columns = config["master_numeric_columns"]

    error_messages = []

    for _, row in master_df.iterrows():
        row_number = row[INTERNAL_MASTER_ROW_NUMBER]

        for column in required_columns:
            if is_blank(row.get(column, "")):
                error_messages.append(f"row {row_number}: {column} is required")

        for column in numeric_columns:
            value = row.get(column, "")

            if is_blank(value):
                continue

            if not can_convert_to_finite_number(value):
                error_messages.append(f"row {row_number}: {column} must be numeric")

    blank_key_rows = [
        str(row[INTERNAL_MASTER_ROW_NUMBER])
        for _, row in master_df.iterrows()
        if is_blank(row.get(master_key, ""))
    ]

    if blank_key_rows:
        error_messages.append(
            f"master_key has blank values at rows: {', '.join(blank_key_rows)}"
        )

    duplicated_key_rows = find_duplicate_master_key_rows(master_df, master_key)

    if duplicated_key_rows:
        error_messages.append(
            "master_key has duplicated values: "
            + "; ".join(duplicated_key_rows)
        )

    if error_messages:
        raise ValueError("Master data is invalid. " + " | ".join(error_messages))

    prepared_df = master_df.copy()

    for column in numeric_columns:
        prepared_df[column] = prepared_df[column].map(format_numeric_value_from_any)

    keep_columns = [master_key] + config["master_value_columns"]
    return prepared_df[keep_columns].copy()


def find_duplicate_master_key_rows(master_df, master_key):
    """Return readable duplicate-key information for master data."""
    duplicated_info = []
    non_blank_df = master_df[~master_df[master_key].map(is_blank)].copy()
    duplicated_mask = non_blank_df.duplicated(subset=[master_key], keep=False)
    duplicated_df = non_blank_df[duplicated_mask]

    for key_value, group_df in duplicated_df.groupby(master_key, sort=True):
        row_numbers = [str(value) for value in group_df[INTERNAL_MASTER_ROW_NUMBER].tolist()]
        duplicated_info.append(f"{key_value} at rows {', '.join(row_numbers)}")

    return duplicated_info


def split_pre_validation_errors(order_df, config):
    """
    Split order rows into rows ready for master matching and validation errors.
    """
    valid_rows = []
    error_rows = []

    for _, row in order_df.iterrows():
        row_dict = row.to_dict()
        error_messages = []

        check_required_values(row_dict, config["order_required_columns"], error_messages)
        normalize_numeric_values(row_dict, config["order_numeric_columns"], error_messages)
        normalize_date_values(
            row_dict,
            config["order_date_columns"],
            config.get("date_format", DEFAULT_DATE_FORMAT),
            error_messages
        )

        if error_messages:
            error_record = create_unmatched_record(
                row_dict,
                config,
                error_type="validation_error",
                error_message="; ".join(error_messages)
            )
            error_rows.append(error_record)
        else:
            valid_rows.append(row_dict)

    valid_df = pd.DataFrame(valid_rows)
    error_df = pd.DataFrame(error_rows)

    return valid_df, error_df


def check_required_values(row_dict, required_columns, error_messages):
    """Add an error when a required value is blank."""
    for column in required_columns:
        if is_blank(row_dict.get(column, "")):
            error_messages.append(f"{column} is required")


def normalize_numeric_values(row_dict, numeric_columns, error_messages):
    """Validate numeric columns and format them for output."""
    for column in numeric_columns:
        value = row_dict.get(column, "")

        if is_blank(value):
            row_dict[column] = ""
            continue

        if not can_convert_to_finite_number(value):
            error_messages.append(f"{column} must be numeric")
            continue

        row_dict[column] = format_numeric_value_from_any(value)


def normalize_date_values(row_dict, date_columns, date_format, error_messages):
    """Validate date columns and format them for output."""
    for column in date_columns:
        value = row_dict.get(column, "")

        if is_blank(value):
            row_dict[column] = ""
            continue

        try:
            parsed_date = pd.to_datetime(value, errors="raise")
        except Exception:
            error_messages.append(f"{column} must be a valid date")
            continue

        row_dict[column] = parsed_date.strftime(date_format)


def normalize_general_value(value):
    """Normalize a general cell value to a safe string value."""
    if is_blank(value):
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    return str(value).strip()


def is_blank(value):
    """Return True when value should be treated as blank."""
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except TypeError:
        pass

    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def can_convert_to_finite_number(value):
    """Return True when value can be converted to a finite number."""
    try:
        numeric_value = pd.to_numeric(value)
        return math.isfinite(float(numeric_value))
    except Exception:
        return False


def format_numeric_value_from_any(value):
    """Convert a numeric-like value to a clean string."""
    numeric_value = float(pd.to_numeric(value))
    return format_numeric_value(numeric_value)


def format_numeric_value(value):
    """Format numeric values without unnecessary .0."""
    numeric_value = float(value)

    if numeric_value.is_integer():
        return str(int(numeric_value))

    return str(numeric_value)


# ------------------------------------------------------------
# Matching and output DataFrame creation
# ------------------------------------------------------------

def match_with_master(valid_order_df, master_df, config):
    """Match valid order rows with master rows by configured keys."""
    order_key = config["order_key"]
    master_key = config["master_key"]

    if len(valid_order_df) == 0:
        return pd.DataFrame(columns=list(valid_order_df.columns) + list(master_df.columns) + [MERGE_STATUS_COLUMN])

    if order_key == master_key:
        return pd.merge(
            valid_order_df,
            master_df,
            how="left",
            on=order_key,
            indicator=True,
            validate="many_to_one"
        )

    return pd.merge(
        valid_order_df,
        master_df,
        how="left",
        left_on=order_key,
        right_on=master_key,
        indicator=True,
        validate="many_to_one"
    )


def build_enriched_df(merged_df, config):
    """Build enriched rows that were successfully matched with master data."""
    if len(merged_df) == 0:
        return pd.DataFrame(columns=config["enriched_output_columns"])

    matched_df = merged_df[merged_df[MERGE_STATUS_COLUMN] == "both"].copy()

    amount_config = config.get("amount_calculation", {"enabled": False})

    if amount_config.get("enabled", False):
        quantity_column = amount_config["quantity_column"]
        price_column = amount_config["price_column"]
        output_column = amount_config["output_column"]

        matched_df[output_column] = [
            format_numeric_value(float(pd.to_numeric(quantity)) * float(pd.to_numeric(price)))
            for quantity, price in zip(matched_df[quantity_column], matched_df[price_column])
        ]

    return select_output_columns(matched_df, config["enriched_output_columns"])


def build_unmatched_df(validation_error_df, merged_df, config):
    """Build unmatched rows from validation errors and master-not-found rows."""
    unmatched_frames = []

    if len(validation_error_df) > 0:
        unmatched_frames.append(validation_error_df)

    if len(merged_df) > 0:
        master_not_found_df = merged_df[merged_df[MERGE_STATUS_COLUMN] == "left_only"].copy()

        if len(master_not_found_df) > 0:
            records = []

            for _, row in master_not_found_df.iterrows():
                records.append(
                    create_unmatched_record(
                        row.to_dict(),
                        config,
                        error_type="master_not_found",
                        error_message=f"{config['order_key']} not found in master"
                    )
                )

            unmatched_frames.append(pd.DataFrame(records))

    if not unmatched_frames:
        return pd.DataFrame(columns=config["unmatched_output_columns"])

    unmatched_df = pd.concat(unmatched_frames, ignore_index=True)
    unmatched_df = select_output_columns(unmatched_df, config["unmatched_output_columns"])

    if "row_number" in unmatched_df.columns:
        unmatched_df = unmatched_df.sort_values(by="row_number", kind="stable")

    return unmatched_df.reset_index(drop=True)


def create_unmatched_record(row_dict, config, error_type, error_message):
    """Create one normalized unmatched output record."""
    record = {}

    for column in config["unmatched_output_columns"]:
        if column == "row_number":
            record[column] = row_dict.get(INTERNAL_ROW_NUMBER, "")
        elif column == "error_type":
            record[column] = error_type
        elif column == "error_message":
            record[column] = error_message
        else:
            record[column] = row_dict.get(column, "")

    return record


def select_output_columns(df, output_columns):
    """Select output columns while creating missing columns as blanks."""
    output_df = df.copy()

    for column in output_columns:
        if column not in output_df.columns:
            output_df[column] = ""

    return output_df[output_columns].copy()


# ------------------------------------------------------------
# Output writing
# ------------------------------------------------------------

def save_enriched_orders(enriched_df, config):
    """Save matched and enriched rows as an Excel file."""
    enriched_path = make_enriched_path(config)

    with pd.ExcelWriter(enriched_path, engine="openpyxl") as writer:
        enriched_df.to_excel(writer, sheet_name="enriched_orders", index=False)
        worksheet = writer.sheets["enriched_orders"]
        apply_basic_worksheet_format(worksheet)

    print(f"Enriched Excel saved: {enriched_path}")


def save_unmatched_orders(unmatched_df, config):
    """Save validation errors and master-not-found rows as an Excel file."""
    unmatched_path = make_unmatched_path(config)

    with pd.ExcelWriter(unmatched_path, engine="openpyxl") as writer:
        unmatched_df.to_excel(writer, sheet_name="unmatched_orders", index=False)
        worksheet = writer.sheets["unmatched_orders"]
        apply_basic_worksheet_format(worksheet)

    print(f"Unmatched Excel saved: {unmatched_path}")


def save_summary(
    total_order_rows,
    master_rows,
    enriched_df,
    unmatched_df,
    validation_error_count,
    master_not_found_count,
    config,
    started_at,
    finished_at
):
    """Save matching summary and settings as an Excel file."""
    summary_path = make_summary_path(config)

    summary_df = pd.DataFrame(
        [
            {"item": "order_file", "value": config["order_file"]},
            {"item": "master_file", "value": config["master_file"]},
            {"item": "enriched_file", "value": config["enriched_file"]},
            {"item": "unmatched_file", "value": config["unmatched_file"]},
            {"item": "summary_file", "value": config["summary_file"]},
            {"item": "total_order_rows", "value": total_order_rows},
            {"item": "matched_rows", "value": len(enriched_df)},
            {"item": "unmatched_rows", "value": len(unmatched_df)},
            {"item": "validation_error_rows", "value": validation_error_count},
            {"item": "master_not_found_rows", "value": master_not_found_count},
            {"item": "started_at", "value": started_at.strftime("%Y-%m-%d %H:%M:%S")},
            {"item": "finished_at", "value": finished_at.strftime("%Y-%m-%d %H:%M:%S")},
        ],
        columns=["item", "value"]
    )

    settings_df = pd.DataFrame(
        [
            {"setting": "order_key", "value": config["order_key"]},
            {"setting": "master_key", "value": config["master_key"]},
            {"setting": "order_required_columns", "value": ", ".join(config["order_required_columns"])},
            {"setting": "master_required_columns", "value": ", ".join(config["master_required_columns"])},
            {"setting": "master_value_columns", "value": ", ".join(config["master_value_columns"])},
            {"setting": "order_numeric_columns", "value": ", ".join(config["order_numeric_columns"])},
            {"setting": "master_numeric_columns", "value": ", ".join(config["master_numeric_columns"])},
            {"setting": "order_date_columns", "value": ", ".join(config["order_date_columns"])},
            {"setting": "date_format", "value": config.get("date_format", DEFAULT_DATE_FORMAT)},
        ],
        columns=["setting", "value"]
    )

    master_check_df = pd.DataFrame(
        [
            {"item": "master_rows", "value": master_rows},
            {"item": "blank_master_keys", "value": 0},
            {"item": "duplicate_master_keys", "value": 0},
        ],
        columns=["item", "value"]
    )

    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        settings_df.to_excel(writer, sheet_name="settings", index=False)
        master_check_df.to_excel(writer, sheet_name="master_check", index=False)

        for worksheet in writer.book.worksheets:
            apply_basic_worksheet_format(worksheet)

    print(f"Summary Excel saved: {summary_path}")


def apply_basic_worksheet_format(worksheet):
    """Apply basic formatting to an openpyxl worksheet."""
    format_header_row(worksheet)
    adjust_column_width(worksheet)
    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions


def format_header_row(worksheet):
    """Apply a simple header style to the first row."""
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def adjust_column_width(worksheet):
    """Adjust column widths based on cell values."""
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


# ------------------------------------------------------------
# Main process
# ------------------------------------------------------------

def main():
    started_at = datetime.now()

    create_sample_files()

    config = load_config()
    validate_config(config)

    order_path = make_order_path(config)
    master_path = make_master_path(config)

    if not os.path.exists(order_path):
        raise FileNotFoundError(f"Order file not found: {order_path}")

    if not os.path.exists(master_path):
        raise FileNotFoundError(f"Master file not found: {master_path}")

    print(f"Order file: {order_path}")
    print(f"Master file: {master_path}")
    print(f"Enriched file: {make_enriched_path(config)}")
    print(f"Unmatched file: {make_unmatched_path(config)}")
    print(f"Summary file: {make_summary_path(config)}")

    order_df = read_table_file(order_path)
    master_df = read_table_file(master_path)

    check_required_columns(order_df, config["order_required_columns"], "order file")
    check_required_columns(master_df, config["master_required_columns"], "master file")

    check_required_columns(
        master_df,
        [config["master_key"]] + config["master_value_columns"],
        "master file"
    )

    total_order_rows = len(order_df)
    master_rows = len(master_df)

    print(f"Order rows found: {total_order_rows}")
    print(f"Master rows found: {master_rows}")

    prepared_order_df = prepare_order_dataframe(order_df)
    prepared_master_df = prepare_master_dataframe(master_df)

    prepared_master_df = validate_and_prepare_master(prepared_master_df, config)

    valid_order_df, validation_error_df = split_pre_validation_errors(prepared_order_df, config)
    merged_df = match_with_master(valid_order_df, prepared_master_df, config)

    enriched_df = build_enriched_df(merged_df, config)
    unmatched_df = build_unmatched_df(validation_error_df, merged_df, config)

    validation_error_count = 0
    master_not_found_count = 0

    if len(unmatched_df) > 0 and "error_type" in unmatched_df.columns:
        validation_error_count = int((unmatched_df["error_type"] == "validation_error").sum())
        master_not_found_count = int((unmatched_df["error_type"] == "master_not_found").sum())

    save_enriched_orders(enriched_df, config)
    save_unmatched_orders(unmatched_df, config)

    finished_at = datetime.now()

    save_summary(
        total_order_rows,
        master_rows,
        enriched_df,
        unmatched_df,
        validation_error_count,
        master_not_found_count,
        config,
        started_at,
        finished_at
    )

    print("Matching finished.")
    print(f"Matched rows: {len(enriched_df)}")
    print(f"Unmatched rows: {len(unmatched_df)}")
    print(f"Validation error rows: {validation_error_count}")
    print(f"Master not found rows: {master_not_found_count}")
    print(f"Output folder: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
